from pathlib import Path
import os
from tkinter import messagebox
import shutil
import xlwings as xw
import shutil
import time
from openpyxl import load_workbook

TEMPLATE_PATH = Path("assets/project_dashboard_template.xlsm")

def create_dashboard_workbook(excel_path: Path) -> bool:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")

    # Archive if exists
    if excel_path.exists():
        try:
            from openpyxl import load_workbook

            # Load and archive sheet
            wb = load_workbook(excel_path, keep_vba=True)
            if "Dashboard" in wb.sheetnames:
                dashboard = wb["Dashboard"]
                archive_name = f"Dashboard_Archive_{time.strftime('%Y%m%d_%H%M%S')}"
                wb.copy_worksheet(dashboard).title = archive_name
                wb.save(excel_path)
                print(f"[INFO] Archived Dashboard to '{archive_name}'")

        except Exception as e:
            print(f"[WARNING] Failed to archive dashboard: {e}")

    # Replace file
    shutil.copy(TEMPLATE_PATH, excel_path)
    return True

def merge_archived_logs(excel_path: Path):
    import xlwings as xw

    app = xw.App(visible=False)
    wb = app.books.open(str(excel_path))

    try:
        # Find archived sheet
        archive_sheet = None
        for sht in wb.sheets:
            if sht.name.startswith("Dashboard_Archive_"):
                archive_sheet = sht
                break

        if not archive_sheet:
            return

        # Read archive data (excluding header)
        data = archive_sheet.range("A2").expand("table").value
        if not data:
            return

        dashboard = wb.sheets["Dashboard"]
        last_row = dashboard.range("A" + str(dashboard.cells.last_cell.row)).end("up").row + 1
        dashboard.range(f"A{last_row}").value = data
        wb.save()
        print(f"[INFO] Merged {len(data)} rows from archive.")

    except Exception as e:
        print(f"[ERROR] Failed to merge archive logs: {e}")

    finally:
        wb.close()
        app.quit()

# if __name__ == "__main__":
#     create_dashboard_workbook(EXCEL_PATH)
